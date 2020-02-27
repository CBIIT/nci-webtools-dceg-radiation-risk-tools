
import os
import requests
import json
import tqdm
import csv
from openpyxl import load_workbook

from django.conf import settings
from django.core.management.base import BaseCommand
from django.http import HttpRequest
from django.http import QueryDict

from irep.calculator.constants import SESSION_FORM_KEY, SUMMARY_REPORT
from irep.calculator.forms import (
    GENDER_CHOICES, CANCER_CHOICES, ETHNIC_CHOICES, EXPOSURE_SOURCE_CHOICES, SMOKING_HISTORY_CHOICES, RADON_EXPRATE_CHOICES, DOSETYPE_CHOICES,
    EXPRATE_CHOICES, RADTYPE_CHOICES,
)
from irep.calculator.views import get_session_parameters

class Command (BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('files', nargs='+')

    def get_default_params_dict(self):
        request = HttpRequest()
        request.session = {}
        postData = QueryDict(query_string='', mutable=True)
        postData.update({'gen_choice': GENDER_CHOICES[2][0], 'by': 1940, 'dod': 2000, 'cancer_choice': CANCER_CHOICES[1][0], })
        postData.update({'ethnic': ETHNIC_CHOICES[0][0], 'exposure_source': EXPOSURE_SOURCE_CHOICES[0][0], 'smoking_history': SMOKING_HISTORY_CHOICES[0][0], })
        postData.update({'radon_-TOTAL_FORMS': 1, 'radon_-INITIAL_FORMS': 0, 'radon_-MAX_NUM_FORMS': 200, 'radon_-0-yoe': 1980, 'radon_-0-exptype': RADON_EXPRATE_CHOICES[1][0], 'radon_-0-dosetype': DOSETYPE_CHOICES[1][0], 'radon_-0-doseparm1': 2, 'radon_-0-doseparm2': 2, 'radon_-0-doseparm3': 0, })
        postData.update({'dose_-TOTAL_FORMS': 1, 'dose_-INITIAL_FORMS': 0, 'dose_-MAX_NUM_FORMS': 200, 'dose_-0-yoe': 1980, 'dose_-0-exprate': EXPRATE_CHOICES[2][0], 'dose_-0-radtype': RADTYPE_CHOICES[5][0], 'dose_-0-dosetype': DOSETYPE_CHOICES[1][0], 'dose_-0-doseparm1': 2, 'dose_-0-doseparm2': 2, 'dose_-0-doseparm3': 0, })
        postData.update({'sample_size': 10000, 'random_seed': 99, 'ududtype': DOSETYPE_CHOICES[1][0], 'ududparm1': 1, 'ududparm2': 1, 'ududparm3': 0, })
        request.session[SESSION_FORM_KEY] = postData
        return get_session_parameters(request)

    def assert_expected_keys(self, response_data):
        if 'sumidx_tab' not in response_data:
            raise Exception("Missing expected key 'sumidx_tab'.")
        if 'summ_tab' not in response_data:
            raise Exception("Missing expected key 'summ_tab'.")
        if 'messages' not in response_data:
            raise Exception("Missing expected key 'messages'.")
    
        for member in ['1st', '2.5th', '5th', '10th', '25th', '50th', '75th', '90th', '95th', '97.5th', '99th']:
            if member not in response_data['sumidx_tab']:
                raise Exception("Missing expected value '%s' in sumidx_tab." % member)

    def get_value_for_label(self, codelist, label):
        for sublist in codelist:
            if sublist[1] == label:
                return sublist[0]
        raise Exception('Could not find value for label "{0}"'.format(label))       

    def handle(self, *args, **options):
        for f in options['files']:
            print('Processing %s' % f)
            wb = load_workbook(filename=f, read_only=True)
            ws = wb['IREP']
            column_count = ws.max_column
            row_count = ws.max_row
            if column_count != 18:
                raise Exception('Expecting 18 columns, got {0}'.format(column_count))

            filename, file_extension = os.path.splitext(f)
            with open(filename + '.csv', "wt") as csv_file:
                writer = csv.writer(csv_file, delimiter=',')
                writer.writerow(['CTB ID', '1st', '2.5th', '5th', '10th', '25th', '50th', '75th', '90th', '95th', '97.5th', '99th', 'ERR'])
                
                # iterate over rows - skipping first, which is the header
                for row in tqdm.tqdm(ws.iter_rows(row_offset=1), total=row_count):
                    # Skip blank records.
                    if not any([c.value for c in row]):
                        continue
                    # get the subject id
                    data = [row[0].value,]
                    # build parameters
                    params = self.get_default_params_dict()
                    params['gen_choice'] = self.get_value_for_label(GENDER_CHOICES, row[1].value)
                    params['by'] = int(row[2].value)
                    params['dod'] = int(row[3].value)
                    params['cancer_choice'] = self.get_value_for_label(CANCER_CHOICES, row[4].value)
                    params['dose_-0-yoe'] = int(row[5].value)
                    params['dose_-0-exprate'] = self.get_value_for_label(EXPRATE_CHOICES, row[6].value)
                    params['dose_-0-radtype'] = self.get_value_for_label(RADTYPE_CHOICES, row[7].value)
                    params['dose_-0-dosetype'] = self.get_value_for_label(DOSETYPE_CHOICES, row[8].value)
                    params['dose_-0-doseparm1'] = row[9].value
                    params['dose_-0-doseparm2'] = row[10].value or 0
                    params['dose_-0-doseparm3'] = row[11].value or 0
                    params['sample_size'] = int(row[12].value)
                    params['random_seed'] = int(row[13].value)
                    params['ududtype'] = self.get_value_for_label(DOSETYPE_CHOICES, row[14].value)
                    params['ududparm1'] = row[15].value
                    params['ududparm2'] = row[16].value or 0
                    params['ududparm3'] = row[17].value or 0
                    # execute analysis
                    r = requests.post('%s%s' % (settings.WINDOWS_SERVER, SUMMARY_REPORT), data=params, headers={'user-agent': 'Django App (%s)' % settings.BASE_URL,})
                    r.raise_for_status()
                    response_data = r.json()
                    # check for expected keys
                    self.assert_expected_keys(response_data)   
                    # write data to csv file
                    data.extend(response_data['summ_tab'])
                    writer.writerow(data)
